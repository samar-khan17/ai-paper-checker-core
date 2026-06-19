# ============================================================
#   core/exporter.py — Excel & PDF export of checked papers
#   Turns graded records into files a teacher can actually use:
#   .xlsx grade sheets and .pdf per-student report cards.
# ============================================================

import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger("exporter")


# ── EXCEL: full grade sheet ───────────────────────────────────
def export_results_to_excel(rows: list, out_path: str, title: str = "Checked Papers") -> str:
    """rows: list of dicts from db.search_submissions(...).
    Writes a formatted .xlsx and returns the path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["#", "Roll No", "Student Name", "Section", "Subject", "Paper",
               "Score", "Out Of", "Percentage", "Grade", "Mode", "Remarks", "Checked On"]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1, value=f"{title} — exported {datetime.now():%Y-%m-%d %H:%M}")
    tcell.font = Font(size=14, bold=True, color="FFFFFF")
    tcell.fill = PatternFill("solid", fgColor="6366F1")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    head_fill = PatternFill("solid", fgColor="13132A")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, r in enumerate(rows, start=1):
        score = r.get("total_score")
        mx = r.get("total_max")
        pct = r.get("percentage")
        vals = [
            i,
            r.get("student_roll", "") or "",
            r.get("student_name", "") or "",
            r.get("class_section", "") or "",
            r.get("subject", "") or "",
            r.get("paper_title", "") or "",
            score if score is not None else "",
            mx if mx is not None else "",
            (f"{pct}%" if pct is not None else ""),
            r.get("grade", "") or "",
            (r.get("checking_mode", "") or "").title(),
            r.get("remarks", "") or "",
            (r.get("submitted_at", "") or "")[:16],
        ]
        rownum = i + 2
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=rownum, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 12 else "left",
                                       vertical="center", wrap_text=(c == 12))
        # Colour the Grade cell by pass/fail
        if pct is not None:
            gcell = ws.cell(row=rownum, column=10)
            gcell.fill = PatternFill("solid",
                                     fgColor=("C6F6D5" if pct >= 50 else "FED7D7"))

    widths = [5, 14, 24, 10, 12, 22, 8, 8, 12, 8, 10, 34, 18]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = w
    ws.freeze_panes = "A3"

    # Summary row
    graded = [r for r in rows if r.get("percentage") is not None]
    if graded:
        avg = round(sum(r["percentage"] for r in graded) / len(graded), 1)
        passed = sum(1 for r in graded if r["percentage"] >= 50)
        srow = len(rows) + 4
        ws.cell(row=srow, column=2, value="SUMMARY").font = Font(bold=True)
        ws.cell(row=srow, column=3, value=f"{len(graded)} graded")
        ws.cell(row=srow + 1, column=3, value=f"Average: {avg}%")
        ws.cell(row=srow + 2, column=3, value=f"Pass (≥50%): {passed}/{len(graded)}")

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(f"Excel exported: {out_path} ({len(rows)} rows)")
    return out_path


# ── PDF: single-student report card ───────────────────────────
def export_student_report_pdf(meta: dict, result: dict, out_path: str) -> str:
    """meta: student/paper info, result: db.get_result_by_submission(...)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    accent = colors.HexColor("#6366f1")
    h = ParagraphStyle("h", parent=styles["Title"], textColor=accent, fontSize=20)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=11,
                         textColor=colors.HexColor("#555555"))
    story = []

    story.append(Paragraph("Smart Paper Checker — Report Card", h))
    story.append(Spacer(1, 4 * mm))

    info = [
        ["Student", meta.get("student_name", "—") or "—",
         "Roll No", meta.get("student_roll", "—") or "—"],
        ["Subject", meta.get("subject", "—") or "—",
         "Paper", meta.get("paper_title", "—") or "—"],
        ["Score", f"{result.get('total_score','-')}/{result.get('total_max','-')}",
         "Grade", result.get("grade", "—")],
        ["Percentage", f"{result.get('percentage','-')}%",
         "Checked On", datetime.now().strftime("%Y-%m-%d")],
    ]
    it = Table(info, colWidths=[28 * mm, 55 * mm, 28 * mm, 55 * mm])
    it.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef0ff")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#eef0ff")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(it)
    story.append(Spacer(1, 6 * mm))

    if result.get("summary_feedback"):
        story.append(Paragraph("<b>Overall Feedback</b>", styles["Heading3"]))
        story.append(Paragraph(result["summary_feedback"], sub))
        story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("<b>Question Breakdown</b>", styles["Heading3"]))
    data = [["Q#", "Score", "Feedback"]]
    qr = result.get("question_results", {})
    for qnum, q in qr.items():
        data.append([str(qnum),
                     f"{q.get('score',0)}/{q.get('max_marks',0)}",
                     Paragraph(str(q.get("feedback", "") or "—"), sub)])
    qt = Table(data, colWidths=[14 * mm, 24 * mm, 128 * mm])
    qt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), accent),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(qt)

    doc.build(story)
    logger.info(f"PDF report exported: {out_path}")
    return out_path
